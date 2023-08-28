from openpyxl.styles import (NamedStyle, Font, Border, Side, PatternFill, Alignment)
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet import worksheet
import re
from excelutils import headers, width_columns, formats


def _try_string_to_number(number: str):
    match number:
        case int() | float() as num:
            return num
        case str() as num if num.isdigit():
            return int(num)
        case str() as num if num.replace(',', '', 1).replace('.', '', 1).isdigit():
            return float(num.replace(",", ".", 1))
        case str() as num if len(num) > 0:
            return num.strip()
        case _:
            return ""


# https://stackoverflow.com/questions/27133731/folding-multiple-rows-with-openpyxl
#
def styles_add(this_book: Workbook):
    name_styles = {}
    for item in formats:
        name_styles[item['name']] = NamedStyle(name=item['name'])
        # header_style = NamedStyle(name=item['name'])
        name_styles[item['name']].font = Font(name=item['font']['name'], bold=item['font']['bold'],
                                              size=item['font']['size'])
        bd = Side(style=item['border']['style'], color=item['border']['color'])
        name_styles[item['name']].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        name_styles[item['name']].fill = PatternFill(patternType=item['fill']['patternType'],
                                                     fgColor=item['fill']['fgColor'])
        name_styles[item['name']].alignment = Alignment(horizontal=item['alignment']['horizontal'],
                                                        vertical=item['alignment']['vertical'],
                                                        wrap_text=item['alignment']['wrap_text'],
                                                        shrink_to_fit=item['alignment']['shrink_to_fit'],
                                                        indent=item['alignment']['indent'])
        if not (item['name'] in this_book.named_styles):
            this_book.add_named_style(name_styles[item['name']])


def create_basic_header(sheet: worksheet):
    sheet.append(["."])
    headers["A:O"].extend(headers["P:T"])
    sheet.append(headers["A:O"])
    for column in range(1, len(headers["A:O"]) + 1):
        sheet.cell(row=2, column=column).style = 'title_basic'
    sheet.cell(row=1, column=column_index_from_string('K')).value = headers['K1']
    sheet.cell(row=1, column=column_index_from_string('K')).style = 'title_basic'
    sheet.merge_cells('K1:M1')

    sheet.cell(row=1, column=column_index_from_string('N')).value = headers['N1']
    sheet.cell(row=1, column=column_index_from_string('N')).style = 'title_basic'
    sheet.merge_cells('N1:O1')

    sheet.cell(row=1, column=column_index_from_string('P')).value = headers['P1']
    sheet.cell(row=1, column=column_index_from_string('P')).style = 'title_basic'
    sheet.merge_cells('P1:T1')

    for width in width_columns:
        sheet.column_dimensions[width].width = width_columns[width]


def _range_decorating(sheet: worksheet, row, columns: list[str], style_name: str):
    """ Устанавливает стиль style_name для ячеек из списка columns """
    for column in columns:
        sheet.cell(row=row, column=column_index_from_string(column)).style = style_name


def _create_table_header(sheet: worksheet, table_info: tuple[str, str, str, str], row: int):
    """ На лист sheet в строку row, информацию table_info """
    tmp = re.findall(r'(^\d+\.\d+)(-\d+)+(-\d+$)', table_info[0])[0]
    extended_name = f"Таблица {tmp[0]}{tmp[2]}. {table_info[1].capitalize()}."

    sheet.cell(row=row, column=column_index_from_string('E')).value = table_info[0]  # код таблицы
    sheet.cell(row=row, column=column_index_from_string('F')).value = ""
    sheet.cell(row=row, column=column_index_from_string('G')).value = extended_name  # table_info[1]  # название таблицы
    sheet.cell(row=row, column=column_index_from_string('H')).value = ""
    sheet.cell(row=row, column=column_index_from_string('I')).value = ""
    sheet.cell(row=row, column=column_index_from_string('J')).value = ""

    _range_decorating(sheet, row, ['E', 'F'], 'line_table')
    _range_decorating(sheet, row-1, ['G', 'H', 'I', 'J'], 'table_name')
    _range_decorating(sheet, row, ['G', 'H', 'I', 'J'], 'table_name')
    # sheet.cell(row=row, column=column_index_from_string('G')).style = 'table_name'

    column = column_index_from_string('K')
    sheet.cell(row=row - 1, column=column).value = headers['K1']
    sheet.cell(row=row - 1, column=column).style = 'further_quotes'
    sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 2)

    sheet.cell(row=row, column=column_index_from_string('K')).value = headers['K']
    sheet.cell(row=row, column=column_index_from_string('L')).value = headers['L']
    sheet.cell(row=row, column=column_index_from_string('M')).value = headers['M']
    _range_decorating(sheet, row, ['K', 'L', 'M'], 'further_quotes')

    column = column_index_from_string('N')
    sheet.cell(row=row - 1, column=column).value = headers['N1']
    sheet.cell(row=row - 1, column=column).style = 'title_attributes'

    attributes = []
    if table_info[2]:
        attributes.extend(table_info[2].split(','))  # атрибуты
        attributes_length = len(attributes)
        for i, attribute in enumerate(attributes):
            sheet.cell(row=row, column=column + i).value = attribute
            sheet.cell(row=row, column=column + i).style = 'title_attributes'

        if attributes_length > 1:
            sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                              end_column=column + attributes_length - 1)
    else:
        sheet.cell(row=row, column=column_index_from_string('N')).value = headers['N']
        sheet.cell(row=row, column=column_index_from_string('O')).value = headers['O']

        sheet.cell(row=row, column=column_index_from_string('N')).style = 'title_attributes'
        sheet.cell(row=row, column=column_index_from_string('O')).style = 'title_attributes'

        sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1, end_column=column + 1)

    parameters = []
    if table_info[3]:
        column += len(attributes)
        parameters.extend(table_info[3].split(','))  # параметры
        parameters_length = len(parameters)
        if parameters_length > 0:
            mini_table_header = headers['P:T']
            parameter_tile_length = len(mini_table_header)  # таблица: 'от', 'до', 'ед.изм.', 'шаг', 'тип'

            for parameter in parameters:
                sheet.cell(row=row - 1, column=column).value = parameter  # название параметра
                sheet.cell(row=row - 1, column=column).style = 'title_parameter'

                i = 0
                for item in mini_table_header:
                    sheet.cell(row=row, column=column + i).value = item
                    sheet.cell(row=row, column=column + i).style = 'title_parameter'
                    i += 1

                column_delta = parameter_tile_length - 1
                sheet.merge_cells(start_row=row - 1, start_column=column, end_row=row - 1,
                                  end_column=column + column_delta)
                column += column_delta + 1
    # sheet.append(["."])


def put_table_to_sheet(sheet: worksheet, table_info: tuple[str, str, str, str], row: int):
    """ На лист sheet в строку row, пишет информацию о таблице table_info.
     Присваивает группу """
    sheet.row_dimensions.group(row, row, outline_level=1)
    _create_table_header(sheet, table_info, row)


def _create_quote_line(sheet: worksheet, quote_info: tuple, row: int):
    """ На лист sheet в строку row, информацию table_info """
    # print(f"в выходной файл на строку {row} выводим расценку {quote_info}")
    # ["idx", "table", "cod", "title", "measure", "stat", "flag", "basic_slave", "link_cod"]

    quote_code_font = Font(name='Calibri', bold=True, size=8, color="000000")
    measure_font = Font(name='Calibri', bold=False, size=8, color="60497A")

    sheet.cell(row=row, column=column_index_from_string('E')).value = quote_info[1]  # код таблицы
    sheet.cell(row=row, column=column_index_from_string('F')).value = quote_info[2]  # код расценки
    sheet.cell(row=row, column=column_index_from_string('G')).value = quote_info[3]  # текст расценки
    sheet.cell(row=row, column=column_index_from_string('H')).value = quote_info[4]  # измеритель
    sheet.cell(row=row, column=column_index_from_string('I')).value = _try_string_to_number(quote_info[5])
    sheet.cell(row=row, column=column_index_from_string('J')).value = '++'
    sheet.cell(row=row, column=column_index_from_string('K')).value = quote_info[7]  # тип расценки
    sheet.cell(row=row, column=column_index_from_string('L')).value = quote_info[8]  # код родительской расценки

    _range_decorating(sheet, row, ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], 'quote_line')
    # sheet.cell(row=row, column=column_index_from_string('I')).font = Font(name='Calibri', bold=False, size=8,
    #                                                                       color="000000")
    sheet.cell(row=row, column=column_index_from_string('I')).alignment = Alignment(horizontal='right')
    sheet.cell(row=row, column=column_index_from_string('J')).alignment = Alignment(horizontal='center')

    sheet.cell(row=row, column=column_index_from_string('F')).font = quote_code_font
    sheet.cell(row=row, column=column_index_from_string('H')).font = measure_font


def put_quote_to_sheet(sheet: worksheet, quote_info: tuple, row: int):
    """ На лист sheet в строку row, пишет информацию о расценке quote_info.
     Присваивает группу """
    sheet.row_dimensions.group(row, row, outline_level=2)
    _create_quote_line(sheet, quote_info, row)


def put_attributes_to_sheet(sheet: worksheet, attributes: list, row: int, table_attributes: str):
    """ На лист sheet в строку row, выводит значение атрибутов.  """
    # column_names = ["quote", "name", "value"]
    if len(attributes) > 0:
        table_attributes_list = [x.strip() for x in table_attributes.split(',')]  # список атрибутов из шапки таблицы
        attribute_titles = {x[2].strip(): x[3].strip() for x in attributes}  # словарь атрибут:значение
        start_column_number = column_index_from_string('N')
        for i, headers_attributes in enumerate(table_attributes_list):
            value_out = attribute_titles.get(headers_attributes, " ")
            sheet.cell(row=row, column=start_column_number + i).value = value_out
            sheet.cell(row=row, column=start_column_number + i).style = 'line_table'


def _get_parameter_from_tuple(src_data: tuple[str, str, str, str, str]) -> dict:
    """ Преобразует кортеж со значениями параметра в словарь """
    # src_data = ("left", "right", "measure", "step", "type")
    p_value = dict()
    p_value["left"], p_value["right"], p_value["measure"], p_value["step"], p_value["type"] = src_data
    return p_value


def put_parameters_to_sheet(sheet: worksheet, parameters: list, row: int, table_parameters: str,
                            attribute_counter: int):
    """ На лист sheet в строку row, выводит значение атрибутов.  """
    # column_names = ("idx", "quote", "name", "left", "right", "measure", "step", "type")
    if len(parameters) > 0:
        table_parameters_list = [x.strip() for x in table_parameters.split(',')]  # список параметров из шапки таблицы
        # делаем словарь параметр: значения
        parameters_dict = {x[2]: _get_parameter_from_tuple(x[3:]) for x in parameters}
        # колонка равна стартовой для атрибутов + кол-во атрибутов
        start_column_number = column_index_from_string('N') + attribute_counter
        step = 0
        # перебираем параметры в шапке таблицы
        for i, headers_parameter in enumerate(table_parameters_list):
            parameter_value = parameters_dict.get(headers_parameter, None)
            if parameter_value:
                column = i + step
                sheet.cell(row=row, column=start_column_number + column).value = _try_string_to_number(
                    parameter_value.get("left", "0"))
                sheet.cell(row=row, column=start_column_number + column + 1).value = _try_string_to_number(
                    parameter_value.get("right", "0"))
                sheet.cell(row=row, column=start_column_number + column + 2).value = _try_string_to_number(
                    parameter_value.get("measure", ""))
                sheet.cell(row=row, column=start_column_number + column + 3).value = _try_string_to_number(
                    parameter_value.get("step", "0"))
                sheet.cell(row=row, column=start_column_number + column + 4).value = _try_string_to_number(
                    parameter_value.get("type", "0"))
                step += 4
                for c in range(i + step + 1):
                    sheet.cell(row=row, column=start_column_number + c).style = 'line_table'
