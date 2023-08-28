import openpyxl
from filesutils import get_full_file_name, does_file_in_use
from colord_setting import console_colors


class ExcelControl:
    def __init__(self, file_name: str = None, file_path: str = None):
        # super().__init__()
        self.file_name = file_name
        self.file_path = file_path
        self.full_name = get_full_file_name(file_name, file_path)
        self.repr_filename = f"'{console_colors['YELLOW']}{self.full_name}{console_colors['RESET']}'"
        self.sheet_name = None
        self.book = None
        self.sheet = None

    def __enter__(self):
        """ Вызывается при старте контекстного менеджера. Открывает книгу, создает листы. """
        self.open_excel_file()
        return self

    def __exit__(self, exception_type, exception_value, traceback):
        """ Будет вызван в завершении конструкции with, или в случае возникновения ошибки после нее. """
        self.close_excel_file()

    def __str__(self):
        return f"excel file: '{self.file_name}', sheet: '{self.sheet_name}', *: '{self.sheet}'"

    def open_excel_file(self):
        """ Открывает файл, если он уже существует или создает новый."""
        if not does_file_in_use(self.full_name):
            try:
                self.book = openpyxl.load_workbook(self.full_name)
            except IOError as err:
                error_out = f"{console_colors['RED']}{err}{console_colors['RESET']}"
                print(f"Ошибка при открытии файла: '{self.full_name}'\n\t{error_out}")
                self.book = openpyxl.Workbook()
                print(f"Создаем новый файл.")
        else:
            print(f"Файл {self.repr_filename} используется другим приложением.")

    def close_excel_file(self):
        if does_file_in_use(self.full_name):
            print(f"Не могу записать файл {self.repr_filename}, используется другим приложением.")
        else:
            if self.book:
                self.book.save(self.full_name)
                self.book.close()

    def delete_all_sheets(self):
        """ Удаляет все листы книги """
        # if self.book:
        #     for sheet in self.book.worksheets:
        #         self.book.remove(sheet)
        #     self.book.create_sheet(".")
        self.book and [self.book.remove(sheet) for sheet in self.book.worksheets]
        self.book and self.book.create_sheet(".")

    def create_sheets(self, sheets_name: tuple[str] = None):

        self.book and sheets_name and [self.book.create_sheet(name) for name in sheets_name]
        if len(self.book.worksheets) > 1 and ("." in self.book.sheetnames):
            self.book.remove(self.book["."])

    def set_sheet_grid(self, grid: bool = True):
        if self.book and self.sheet:
            self.sheet.sheet_view.showGridLines = grid
